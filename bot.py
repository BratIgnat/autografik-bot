import os
from datetime import datetime, timedelta
from uuid import uuid4
import io

from aiogram import Bot, Dispatcher, types, F
from aiogram.filters import Command
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import StatesGroup, State
from aiogram.types import (
    ReplyKeyboardMarkup,
    KeyboardButton,
    ReplyKeyboardRemove,
    FSInputFile,
    CallbackQuery,
)
from aiogram.utils.keyboard import InlineKeyboardBuilder
from dotenv import load_dotenv
from supabase import create_client, Client

# --- matplotlib для сервера без дисплея ---
import matplotlib
matplotlib.use("Agg")            # важно для ВМ
import matplotlib.pyplot as plt  # после выбора бэкенда

# --- openpyxl для экспорта в Excel ---
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------- ENV & INIT ----------------
load_dotenv()

SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_KEY = os.getenv("SUPABASE_KEY")
TELEGRAM_TOKEN = os.getenv("TELEGRAM_TOKEN") or os.getenv("BOT_TOKEN")

if not TELEGRAM_TOKEN:
    print("ERROR: TELEGRAM_TOKEN (или BOT_TOKEN) не найден в .env")
    raise SystemExit(1)
print("TOKEN OK: ****" + TELEGRAM_TOKEN[-6:])

if not SUPABASE_URL or not SUPABASE_KEY:
    print("WARN: SUPABASE_URL/SUPABASE_KEY не заданы — проверь .env")

supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)
bot = Bot(token=TELEGRAM_TOKEN)
dp = Dispatcher()

# ---------------- STATES ----------------
class SlotState(StatesGroup):
    waiting_for_date = State()
    waiting_for_slot = State()

class JoinTeamState(StatesGroup):
    waiting_for_invite = State()

class CreateTeamState(StatesGroup):
    waiting_for_team_name = State()

# Админ-панель: недели и лимиты
class AdminWeekState(StatesGroup):
    waiting_for_monday = State()

class AdminLimitsState(StatesGroup):
    choosing_date = State()
    choosing_scope = State()   # day|slot
    choosing_slot = State()    # если scope=slot
    choosing_role = State()
    waiting_for_count = State()

# Админ-панель: участники
class AdminMembersState(StatesGroup):
    browsing = State()
    member_card = State()
    changing_role = State()

# ---------------- CONSTS & HELPERS ----------------
ROLE_HEADERS = {'Официанты','Бармен','Хостес','Ранеры','Админы','Стажёры','Другие'}
ROLE_CODES = [
    ("Официанты", "employee"),
    ("Хостес",     "host"),
    ("Бармен",     "barman"),
    ("Ранеры",     "runner"),
    ("Админы",     "admin"),
    ("Стажёры",    "trainee"),
]
STD_SLOTS = ["09:30-23:00", "10:00-23:00", "11:00-23:00", "12:00-23:00", "13:00-23:00", "17:00-23:00"]
NO_SHIFT = {"-", "вых", "выходной"}  # значения, не считающиеся сменой
PAGE_SIZE = 10

# Цвета и стиль Excel
XL_GROUP_FILL = "FFA500"   # оранжевый для шапок подгрупп
XL_HEADER_FILL = "EFEFEF"  # серый для заголовков таблиц
XL_WEEKEND_FILL = "FFF2CC" # мяглый фон для сб/вс
XL_BORDER = Side(style="thin", color="DDDDDD")


def ensure_admin(user_row: dict) -> bool:
    return bool(user_row and (user_row.get("is_admin") or user_row.get("is_owner")))


def now_iso_z() -> str:
    # UTC ISO8601 с Z — нормально пишется в timestamptz
    return datetime.utcnow().isoformat() + "Z"


def is_cancel(text: str) -> bool:
    """Возвращает True, если сообщение похоже на 'Отмена' (с эмодзи/пробелами/регистром)."""
    return bool(text) and ("отмена" in text.casefold())

# ---------------- KEYBOARDS ----------------

def menu_keyboard():
    kb = [
        [KeyboardButton(text="📅 Расписание"), KeyboardButton(text="📝 Моя смена")],
        [KeyboardButton(text="🖼 Поменять фон"), KeyboardButton(text="👥 Пригласить сотрудника")],
        [KeyboardButton(text="👤 Выдать роль"), KeyboardButton(text="❓ Помощь")],
    ]
    return ReplyKeyboardMarkup(keyboard=kb, resize_keyboard=True)


def start_keyboard():
    kb = [
        [KeyboardButton(text="➕ Создать команду")],
        [KeyboardButton(text="🔑 Вступить по коду")]
    ]
    return ReplyKeyboardMarkup(keyboard=kb, resize_keyboard=True)


def cancel_kb() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        keyboard=[[KeyboardButton(text="❌ Отмена")]],
        resize_keyboard=True
    )

# ---------------- DATA HELPERS ----------------

def get_active_week(team_id):
    week_resp = supabase.table("weeks").select("*").eq("team_id", team_id).eq("is_active", True).execute()
    if not week_resp.data:
        return None
    return week_resp.data[0]


def get_week_dates(start_date, end_date):
    wdays = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"]
    dates = []
    d0 = datetime.strptime(start_date, "%Y-%m-%d")
    d1 = datetime.strptime(end_date, "%Y-%m-%d")
    for i in range((d1 - d0).days + 1):
        cur = d0 + timedelta(days=i)
        dates.append({
            "weekday": wdays[cur.weekday()],
            "date": cur.strftime("%d.%m"),
            "date_iso": cur.strftime("%Y-%m-%d"),
        })
    return dates

# ---------------- SCHEDULE RENDER (IMAGE) ----------------

def make_schedule_image(users, week_days, shifts):
    # Показываем только активных (если поля нет — считаем активным)
    users = [u for u in users if u.get("is_active", True)]

    columns = ["ФИО"] + [f"{day['weekday']} {day['date']}" for day in week_days]
    role_map = {
        "employee": "Официанты",
        "barman": "Бармен",
        "host": "Хостес",
        "runner": "Ранеры",
        "admin": "Админы",
        "trainee": "Стажёры"
    }
    roles_order = ["employee", "barman", "host", "runner", "admin", "trainee", "other"]
    data_rows = []
    header_rows = []

    for role in roles_order:
        if role == "other":
            role_users = [u for u in users if not u.get('role') or u.get('role') not in roles_order[:-1]]
            if not role_users:
                continue
            data_rows.append(["Другие"] + [""] * len(week_days))
            header_rows.append(len(data_rows)-1)
        else:
            role_users = [u for u in users if u.get('role') == role]
            if not role_users:
                continue
            data_rows.append([role_map[role]] + [""] * len(week_days))
            header_rows.append(len(data_rows)-1)
        for u in role_users:
            row = [u["name"]]
            for day in week_days:
                slot = next((s["slot"] for s in shifts if s["user_id"] == u["id"] and s["date"] == day["date_iso"]), "-")
                row.append(slot)
            data_rows.append(row)

    n_cols = len(columns)
    n_rows = len(data_rows)
    fig_w = min(max(2 + n_cols * 1.35, 8), 24)
    fig_h = min(max(1.8 + n_rows * 0.7, 3), 28)
    fig, ax = plt.subplots(figsize=(fig_w, fig_h))
    ax.axis('off')
    table = ax.table(cellText=data_rows, colLabels=columns, cellLoc='center', loc='center', bbox=[0,0,1,1])
    table.auto_set_font_size(False)
    table.set_fontsize(13)
    table.auto_set_column_width(col=list(range(n_cols)))

    for (row, col), cell in table.get_celld().items():
        if row == 0:
            cell.set_fontsize(14)
            cell.set_text_props(weight="bold")
            cell.set_facecolor("#e3ebfa")
        elif col == 0 and row > 0 and data_rows[row - 1][0] in ROLE_HEADERS:
            cell.set_facecolor("#FFD580")
            cell.set_text_props(weight="bold", color="black")
        else:
            cell.set_facecolor("white")
            cell.set_text_props(weight="normal", color="black")

    plt.tight_layout()
    plt.savefig("schedule.png", bbox_inches='tight', transparent=True, dpi=170)
    plt.close(fig)
    return "schedule.png"

# ---------------- EXCEL EXPORT ----------------

ROLE_TITLE_BY_CODE = {
    "employee": "Официанты",
    "host": "Хостес",
    "barman": "Бармен",
    "runner": "Ранеры",
    "admin": "Админы",
    "trainee": "Стажёры",
}

ROLES_EXPORT_ORDER = ["employee", "host", "barman", "runner", "admin", "trainee", "other"]


def _xl_apply_table_styles(ws, start_row: int, end_row: int, start_col: int, end_col: int, weekend_cols: list[int]):
    # границы + выравнивание
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
            cell.border = Border(top=XL_BORDER, bottom=XL_BORDER, left=XL_BORDER, right=XL_BORDER)

    # подсветка выходных в содержимом (не шапке)
    for c in weekend_cols:
        for r in range(start_row + 1, end_row + 1):
            ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=XL_WEEKEND_FILL)


def _xl_auto_width(ws):
    for col in range(1, ws.max_column + 1):
        max_len = 0
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(col)].width = min(max(12, max_len + 2), 38)


def export_schedule_excel_for_team(team_id: str):
    """
    Возвращает (filename: str, bytes_io: io.BytesIO) с Excel-графиком активной недели.
    Ничего не меняет в БД. Предполагает, что неделя уже активна.
    """
    week = get_active_week(team_id)
    if not week:
        raise RuntimeError("Нет активной недели")

    week_days = get_week_dates(week["start_date"], week["end_date"])  # список из 7 dict

    # Пользователи (только активные)
    users = supabase.table("users").select("id,name,role,is_active").eq("team_id", team_id).order("name").execute().data
    users = [u for u in users if u.get("is_active", True)]

    # Смены только в диапазоне недели
    start_iso = week_days[0]["date_iso"]
    end_iso = week_days[-1]["date_iso"]
    shifts = supabase.table("shifts").select("user_id,date,slot").eq("team_id", team_id) \
        .gte("date", start_iso).lte("date", end_iso).execute().data
    shift_map = {(s["user_id"], s["date"]): (s.get("slot") or "") for s in shifts}

    # Лимиты недели (для сводки наверху)
    limits = supabase.table("limits").select("date,slot,role,max_count").eq("team_id", team_id) \
        .gte("date", start_iso).lte("date", end_iso).execute().data

    # Построение Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "График"

    # Заголовок
    title = f"График на неделю: {week['start_date']} — {week['end_date']}"
    ws.append([title])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=1 + len(week_days))
    ws.cell(row=1, column=1).font = Font(bold=True)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="left", vertical="center")

    cur_row = 3  # отступ после заголовка

    # Сводка лимитов
    ws.cell(row=cur_row, column=1, value="День")
    ws.cell(row=cur_row, column=2, value="Роль")
    ws.cell(row=cur_row, column=3, value="Лимит")
    for c in range(1, 4):
        cell = ws.cell(row=cur_row, column=c)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor=XL_HEADER_FILL)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    cur_row += 1

    # агрегируем лимиты по датам
    day_limits = {}
    for li in limits:
        day_limits.setdefault(li["date"], []).append(li)

    if not day_limits:
        ws.append(["—", "—", "—"])
        cur_row += 1
    else:
        # идём по дням недели по порядку
        for d in week_days:
            daily = day_limits.get(d["date_iso"], [])
            if not daily:
                ws.append([f"{d['weekday']} {d['date']}", "—", "—"])
                cur_row += 1
                continue
            for li in sorted(daily, key=lambda x: (x.get("role") or "", x.get("slot") or "")):
                role = li.get("role") or "—"
                lim = li.get("max_count")
                ws.append([f"{d['weekday']} {d['date']}", role, lim])
                cur_row += 1

    cur_row += 2  # отступ перед блоками по ролям

    # Заголовки колонок для табличной части
    col_labels = ["Сотрудник"] + [d["date"] for d in week_days]

    # Определим индексы колонок выходных (Сб/Вс)
    weekend_cols = []
    for idx, d in enumerate(week_days, start=2):  # начинаем с 2, т.к. 1-й столбец — ФИО
        if d["weekday"] in ("Сб", "Вс"):
            weekend_cols.append(idx)

    # Группировка пользователей по ролям
    users_by_role: dict[str, list[dict]] = {}
    for u in users:
        code = (u.get("role") or "").strip()
        if code not in ROLE_TITLE_BY_CODE:
            code = "other"
        users_by_role.setdefault(code, []).append(u)

    # Сортировка по имени внутри ролей
    for lst in users_by_role.values():
        lst.sort(key=lambda x: (x.get("name") or "").lower())

    # Рендер блоков по ролям
    for role_code in ROLES_EXPORT_ORDER:
        role_users = users_by_role.get(role_code, [])
        if not role_users:
            continue

        # Строка-заголовок подгруппы (оранжевая)
        ws.insert_rows(cur_row, 1)
        ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=len(col_labels))
        gcell = ws.cell(row=cur_row, column=1)
        gcell.value = ROLE_TITLE_BY_CODE.get(role_code, "Другие")
        gcell.font = Font(bold=True)
        gcell.alignment = Alignment(horizontal="left", vertical="center")
        gcell.fill = PatternFill("solid", fgColor=XL_GROUP_FILL)
        cur_row += 1

        # Шапка таблицы
        for c, label in enumerate(col_labels, start=1):
            cell = ws.cell(row=cur_row, column=c, value=label)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor=XL_HEADER_FILL)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        header_row = cur_row
        cur_row += 1

        # Строки сотрудников
        start_table_row = header_row
        for u in role_users:
            ws.cell(row=cur_row, column=1, value=u["name"])  # ФИО — оставляем чёрным для печати
            # ячейки на 7 дней
            for idx, d in enumerate(week_days, start=2):
                slot = shift_map.get((u["id"], d["date_iso"]), "")
                ws.cell(row=cur_row, column=idx, value=slot)
            cur_row += 1
        end_table_row = cur_row - 1

        # Стили сетки и выходных
        _xl_apply_table_styles(ws, start_table_row, end_table_row, 1, len(col_labels), weekend_cols)

        cur_row += 2  # отступ перед следующей подгруппой

    # Автоширина
    _xl_auto_width(ws)

    # Сохранение в BytesIO
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    filename = f"grafik_{week['start_date']}_{week['end_date']}.xlsx"
    return filename, bio

# ---------------- COMMANDS ----------------

@dp.message(Command("start"))
async def cmd_start(message: types.Message, state: FSMContext):
    user = supabase.table("users").select("*").eq("telegram_id", message.from_user.id).execute().data
    if not user:
        supabase.table("users").insert({
            "telegram_id": message.from_user.id,
            "name": message.from_user.full_name or message.from_user.username or f"user_{message.from_user.id}",
            "team_id": None,
            "is_owner": False,
            "is_admin": False,
            "role": None,
            "is_active": True
        }).execute()
        user = [{"team_id": None, "is_active": True}]

    u = user[0]
    if u.get("team_id"):
        if not u.get("is_active", True):
            await message.answer("Твой профиль в команде отключён. Обратись к администратору.")
            return
        await message.answer("Добро пожаловать! Используй меню или команды для работы с ботом.", reply_markup=menu_keyboard())
    else:
        await message.answer("Ты не в команде! Создай команду или вступи по коду:", reply_markup=start_keyboard())


@dp.message(F.text == "➕ Создать команду")
async def btn_create_team(message: types.Message, state: FSMContext):
    await message.answer("Введи название для своей команды (одной строкой):", reply_markup=cancel_kb())
    await state.set_state(CreateTeamState.waiting_for_team_name)


@dp.message(CreateTeamState.waiting_for_team_name)
async def create_team_name(message: types.Message, state: FSMContext):
    if is_cancel(message.text):
        await state.clear()
        await message.answer("❌ Отменено.", reply_markup=menu_keyboard())
        return

    name = message.text.strip()
    invite_code = str(uuid4()).split('-')[0].upper()
    team_id = str(uuid4())
    supabase.table('teams').insert({
        "id": team_id,
        "name": name,
        "invite_code": invite_code,
    }).execute()
    supabase.table('users').update({
        "team_id": team_id,
        "is_owner": True,
        "is_admin": True,
        "is_active": True
    }).eq('telegram_id', message.from_user.id).execute()
    await message.answer(
        f"Команда <b>{name}</b> создана!\nТвой код для приглашения: <code>{invite_code}</code>\nТы назначен владельцем и администратором.",
        parse_mode="HTML",
        reply_markup=menu_keyboard()
    )
    await state.clear()


@dp.message(F.text == "🔑 Вступить по коду")
async def btn_join_team(message: types.Message, state: FSMContext):
    await message.answer("Введи код приглашения (invite_code) команды:", reply_markup=cancel_kb())
    await state.set_state(JoinTeamState.waiting_for_invite)


@dp.message(JoinTeamState.waiting_for_invite)
async def join_team_code(message: types.Message, state: FSMContext):
    if is_cancel(message.text):
        await state.clear()
        await message.answer("❌ Отменено.", reply_markup=menu_keyboard())
        return

    code = message.text.strip().upper()
    team = supabase.table('teams').select('id', 'name').eq('invite_code', code).execute().data
    if not team:
        await message.answer("Команда с таким кодом не найдена. Проверь правильность кода и попробуй снова.")
        return
    team_id = team[0]['id']
    update_result = supabase.table('users').update({
        "team_id": team_id,
        "is_owner": False,
        "is_admin": False,
        "is_active": True
    }).eq('telegram_id', message.from_user.id).execute()
    if update_result.count == 0:
        supabase.table('users').insert({
            "telegram_id": message.from_user.id,
            "name": message.from_user.full_name or message.from_user.username or f"user_{message.from_user.id}",
            "team_id": team_id,
            "is_owner": False,
            "is_admin": False,
            "role": None,
            "is_active": True
        }).execute()
    await message.answer(
        f"Ты успешно вступил в команду <b>{team[0]['name']}</b>!",
        parse_mode="HTML",
        reply_markup=menu_keyboard()
    )
    await state.clear()


@dp.message(F.text == "📅 Расписание")
async def btn_schedule(message: types.Message, state: FSMContext):
    user_resp = supabase.table("users").select("*").eq("telegram_id", message.from_user.id).execute()
    if not user_resp.data or not user_resp.data[0].get("team_id"):
        await message.answer("Ты не состоишь ни в одной команде.")
        return
    if not user_resp.data[0].get("is_active", True):
        await message.answer("Твой профиль в команде отключён. Обратись к администратору.")
        return

    team_id = user_resp.data[0]["team_id"]
    week = get_active_week(team_id)
    if not week:
        await message.answer("Нет активной недели. Пусть владелец команды её создаст.")
        return
    week_days = get_week_dates(week["start_date"], week["end_date"])
    users = supabase.table("users").select("id,name,role,is_active").eq("team_id", team_id).execute().data
    shifts = supabase.table("shifts").select("*").eq("team_id", team_id).execute().data
    img_path = make_schedule_image(users, week_days, shifts)
    photo = FSInputFile(img_path)
    await message.answer_photo(photo, caption="Текущее расписание:", reply_markup=menu_keyboard())


@dp.message(F.text == "👥 Пригласить сотрудника")
async def btn_invite(message: types.Message, state: FSMContext):
    user = supabase.table("users").select("team_id").eq("telegram_id", message.from_user.id).execute().data
    if not user or not user[0].get("team_id"):
        await message.answer("Ты не состоишь ни в одной команде.")
        return
    team_id = user[0]["team_id"]
    team = supabase.table("teams").select("invite_code").eq("id", team_id).execute().data
    invite_code = team[0]["invite_code"] if team and team[0].get("invite_code") else "Нет кода"
    await message.answer(f"Код приглашения для вашей команды: <code>{invite_code}</code>", parse_mode="HTML")


@dp.message(F.text == "📝 Моя смена")
async def myslot_start(message: types.Message, state: FSMContext):
    user = supabase.table("users").select("*").eq("telegram_id", message.from_user.id).execute().data
    if not user or not user[0].get("team_id"):
        await message.answer("Ты не состоишь ни в одной команде.")
        return
    if not user[0].get("is_active", True):
        await message.answer("Твой профиль в команде отключён. Обратись к администратору.")
        return

    team_id = user[0]["team_id"]
    week = get_active_week(team_id)
    if not week:
        await message.answer("Нет активной недели для выбора смены.")
        return
    week_days = get_week_dates(week["start_date"], week["end_date"])

    # Клавиатура с датами + Отмена
    kb_rows = [[KeyboardButton(text=f"{day['weekday']} {day['date']}")] for day in week_days]
    kb_rows.append([KeyboardButton(text="❌ Отмена")])
    await message.answer("Выбери день для смены:", reply_markup=ReplyKeyboardMarkup(keyboard=kb_rows, resize_keyboard=True))
    await state.set_state(SlotState.waiting_for_date)
    await state.update_data(week_days=week_days, team_id=team_id)


@dp.message(F.text == "👤 Выдать роль")
async def btn_give_role(message: types.Message, state: FSMContext):
    user_id = message.from_user.id
    user = supabase.table('users').select('id, team_id, is_admin, is_owner').eq('telegram_id', user_id).execute().data
    if not user or not (user[0].get('is_admin') or user[0].get('is_owner')):
        await message.answer("Только админ или владелец команды может выдавать роли.")
        return

    team_id = user[0]['team_id']
    members = supabase.table('users').select('id, name, role').eq('team_id', team_id).execute().data
    if not members:
        await message.answer("В команде нет сотрудников.")
        return

    keyboard = InlineKeyboardBuilder()
    for member in members:
        keyboard.button(
            text=f"{member['name']} ({member.get('role', '')})",
            callback_data=f"setrole_{member['id']}"
        )
    await message.answer("Выберите сотрудника для назначения роли:", reply_markup=keyboard.as_markup())


@dp.callback_query(F.data.startswith("setrole_"))
async def callback_choose_role(call: CallbackQuery, state: FSMContext):
    member_id = call.data.replace("setrole_", "")
    await state.update_data(member_id=member_id)
    roles = [
        ("ОФИЦИАНТ", "employee"),
        ("ХОСТ", "host"),
        ("БАРМЕН", "barman"),
        ("РАНЕР", "runner"),
        ("АДМИН", "admin"),
        ("СТАЖЁР", "trainee"),
    ]
    keyboard = InlineKeyboardBuilder()
    for title, code in roles:
        keyboard.button(
            text=title,
            callback_data=f"setroleto_{code}"
        )
    keyboard.adjust(2)
    await call.message.edit_text("Выберите новую роль для сотрудника:", reply_markup=keyboard.as_markup())
    await call.answer()


@dp.callback_query(F.data.startswith("setroleto_"))
async def callback_set_role(call: CallbackQuery, state: FSMContext):
    role_code = call.data.replace("setroleto_", "")
    data = await state.get_data()
    member_id = data.get("member_id")
    if not member_id:
        await call.answer("Ошибка: не выбран сотрудник.", show_alert=True)
        return

    supabase.table('users').update({'role': role_code}).eq('id', member_id).execute()
    await call.message.edit_text(f"Роль успешно обновлена!")
    await call.answer("Роль назначена.", show_alert=True)

# ---------------- SLOT PICKING ----------------

@dp.message(SlotState.waiting_for_date)
async def slot_choose_day(message: types.Message, state: FSMContext):
    # ранняя отмена
    if is_cancel(message.text):
        await state.clear()
        await message.answer("❌ Отменено.", reply_markup=menu_keyboard())
        return

    data = await state.get_data()
    week_days = data.get("week_days")
    selected = message.text

    day = next((d for d in week_days if f"{d['weekday']} {d['date']}" == selected), None)
    if not day:
        await message.answer("Неверная дата. Попробуй ещё раз.")
        return

    # клавиатура слотов + отмена
    kb_rows = [[KeyboardButton(text=s)] for s in STD_SLOTS] + [[KeyboardButton(text="вых")], [KeyboardButton(text="❌ Отмена")]]
    await state.update_data(selected_date=day["date_iso"])
    await message.answer("Выбери смену:", reply_markup=ReplyKeyboardMarkup(keyboard=kb_rows, resize_keyboard=True))
    await state.set_state(SlotState.waiting_for_slot)


@dp.message(SlotState.waiting_for_slot)
async def slot_choose_slot(message: types.Message, state: FSMContext):
    # ранняя отмена
    if is_cancel(message.text):
        await state.clear()
        await message.answer("❌ Отменено.", reply_markup=menu_keyboard())
        return

    data = await state.get_data()
    slot = (message.text or "").strip()

    user = supabase.table("users").select("id,team_id,role,is_active").eq("telegram_id", message.from_user.id).execute().data[0]
    if not user.get("is_active", True):
        await message.answer("Твой профиль в команде отключён. Обратись к администратору.")
        await state.clear()
        return

    user_id = user["id"]
    team_id = user["team_id"]
    role    = user["role"]
    date    = data["selected_date"]  # YYYY-MM-DD

    # Если пользователь выбирает "выходной" — пропускаем лимиты
    if slot in NO_SHIFT:
        existing = supabase.table("shifts").select("id").eq("user_id", user_id) \
            .eq("date", date).eq("team_id", team_id).execute().data
        if existing:
            supabase.table("shifts").update({"slot": slot}).eq("id", existing[0]["id"]).execute()
        else:
            supabase.table("shifts").insert({
                "user_id": user_id, "team_id": team_id, "date": date, "slot": slot
            }).execute()
        await message.answer(f"✅ Готово! Ты поставил {slot!r} на {date}.", reply_markup=menu_keyboard())
        await btn_schedule(message, state)
        await state.clear()
        return

    # --- 1) берём все лимиты на этот день и роль (и слот, и дневные) ---
    lim_rows = supabase.table("limits").select("slot,max_count") \
        .eq("team_id", team_id).eq("date", date).eq("role", role).execute().data

    # Выбираем применимый лимит: приоритет точного слота, иначе дневной (slot NULL)
    max_count = None
    limit_is_daily = False
    for r in lim_rows:
        if r["slot"] == slot:
            max_count = r["max_count"]
            limit_is_daily = False
            break
    if max_count is None:
        for r in lim_rows:
            if r["slot"] is None:
                max_count = r["max_count"]
                limit_is_daily = True
                break

    # --- 2) если лимит задан — проверяем занятость ---
    if max_count is not None:
        if limit_is_daily:
            taken_resp = supabase.table("shifts").select("user_id,slot") \
                .eq("team_id", team_id).eq("date", date).execute().data
            taken_ids = [r["user_id"] for r in taken_resp
                         if r["user_id"] != user_id and (r["slot"] or "").strip() not in NO_SHIFT]
        else:
            taken_resp = supabase.table("shifts").select("user_id") \
                .eq("team_id", team_id).eq("date", date).eq("slot", slot).execute().data
            taken_ids = [r["user_id"] for r in taken_resp if r["user_id"] != user_id]

        current_role_count = 0
        if taken_ids:
            roles_resp = supabase.table("users").select("id,role,is_active").in_("id", taken_ids).execute().data
            current_role_count = sum(1 for u in roles_resp if u["role"] == role and u.get("is_active", True))

        if current_role_count >= max_count:
            await message.answer(
                f"🚫 Лимит для роли «{role}» на {date} "
                f"{'(на весь день)' if limit_is_daily else f'в слоте {slot}'} исчерпан: "
                f"{current_role_count}/{max_count}. Выбери другой слот или день.",
                reply_markup=menu_keyboard()
            )
            await state.clear()
            return

    existing = supabase.table("shifts").select("id") \
        .eq("user_id", user_id).eq("date", date).eq("team_id", team_id).execute().data
    if existing:
        supabase.table("shifts").update({"slot": slot}).eq("id", existing[0]["id"]).execute()
    else:
        supabase.table("shifts").insert({
            "user_id": user_id, "team_id": team_id, "date": date, "slot": slot
        }).execute()

    await message.answer(f"✅ Готово! Ты выбрал смену {slot} на {date}.", reply_markup=menu_keyboard())
    await btn_schedule(message, state)
    await state.clear()

# ---------------- ADMIN PANEL ----------------

@dp.message(Command("admin"))
async def admin_entry(message: types.Message, state: FSMContext):
    me = supabase.table("users").select("id,team_id,is_admin,is_owner").eq("telegram_id", message.from_user.id).execute().data
    if not me or not ensure_admin(me[0]):
        await message.answer("Доступ только для админов/владельцев.")
        return
    kb = InlineKeyboardBuilder()
    kb.button(text="📆 Активная неделя", callback_data="admin_week")
    kb.button(text="📈 Лимиты (создать/изменить)", callback_data="admin_limits")
    kb.button(text="👀 Лимиты недели (просмотр)", callback_data="admin_limits_view")
    kb.button(text="🔁 Скопировать лимиты → след. неделя", callback_data="admin_limits_copy_next")
    kb.button(text="📥 Скачать Excel", callback_data="admin_download_excel")
    kb.button(text="👤 Участники", callback_data="admin_members")
    kb.button(text="♻️ Сбросить инвайт-код", callback_data="admin_reset_invite")
    kb.adjust(1)
    await message.answer("Админ-панель:", reply_markup=kb.as_markup())


# --- Active Week flow ---
@dp.callback_query(F.data == "admin_week")
async def admin_week_start(call: CallbackQuery, state: FSMContext):
    me = supabase.table("users").select("team_id,is_admin,is_owner").eq("telegram_id", call.from_user.id).execute().data
    if not me or not ensure_admin(me[0]):
        await call.answer("Нет доступа", show_alert=True); return
    await state.update_data(team_id=me[0]["team_id"])
    txt = ("Введи дату ПОНЕДЕЛЬНИКА в формате YYYY-MM-DD.\n"
           "Я поставлю конец недели = +6 дней и сделаю её активной.")
    await call.message.edit_text(txt)
    # показать кнопку Отмена
    await call.message.answer("Можно отменить ввод:", reply_markup=cancel_kb())
    await state.set_state(AdminWeekState.waiting_for_monday)
    await call.answer()


@dp.message(AdminWeekState.waiting_for_monday)
async def admin_week_set(message: types.Message, state: FSMContext):
    if is_cancel(message.text):
        await state.clear()
        await message.answer("❌ Отменено.", reply_markup=menu_keyboard())
        return

    try:
        dt = datetime.strptime(message.text.strip(), "%Y-%m-%d").date()
    except Exception:
        await message.answer("Неверный формат. Пример: 2025-08-18"); return
    monday = dt - timedelta(days=dt.weekday())
    sunday = monday + timedelta(days=6)

    data = await state.get_data()
    team_id = data["team_id"]

    supabase.table("weeks").update({"is_active": False}).eq("team_id", team_id).eq("is_active", True).execute()
    supabase.table("weeks").insert({
        "team_id": team_id,
        "start_date": monday.isoformat(),
        "end_date": sunday.isoformat(),
        "is_active": True
    }).execute()

    await message.answer(f"✅ Неделя {monday} — {sunday} установлена активной.", reply_markup=menu_keyboard())
    await state.clear()


# --- Limits flow: создание/изменение ---
@dp.callback_query(F.data == "admin_limits")
async def admin_limits_start(call: CallbackQuery, state: FSMContext):
    me = supabase.table("users").select("team_id,is_admin,is_owner").eq("telegram_id", call.from_user.id).execute().data
    if not me or not ensure_admin(me[0]):
        await call.answer("Нет доступа", show_alert=True); return

    team_id = me[0]["team_id"]
    week = get_active_week(team_id)
    if not week:
        await call.message.edit_text("Сначала создай активную неделю (меню → 📆 Активная неделя).")
        await call.answer(); return

    days = get_week_dates(week["start_date"], week["end_date"])
    kb = InlineKeyboardBuilder()
    for d in days:
        kb.button(text=f"{d['weekday']} {d['date']}", callback_data=f"limit_date:{d['date_iso']}")
    kb.adjust(3)
    await state.update_data(team_id=team_id)
    await call.message.edit_text("Выбери день для лимита:", reply_markup=kb.as_markup())
    await state.set_state(AdminLimitsState.choosing_date)
    await call.answer()


@dp.callback_query(AdminLimitsState.choosing_date, F.data.startswith("limit_date:"))
async def admin_limits_pick_date(call: CallbackQuery, state: FSMContext):
    date_iso = call.data.split(":",1)[1]
    await state.update_data(date=date_iso)
    kb = InlineKeyboardBuilder()
    kb.button(text="Лимит на ДЕНЬ", callback_data="limit_scope:day")
    kb.button(text="Лимит на СЛОТ", callback_data="limit_scope:slot")
    kb.adjust(1)
    await call.message.edit_text(f"Дата: {date_iso}\nВыбери тип лимита:", reply_markup=kb.as_markup())
    await state.set_state(AdminLimitsState.choosing_scope)
    await call.answer()


@dp.callback_query(AdminLimitsState.choosing_scope, F.data.startswith("limit_scope:"))
async def admin_limits_pick_scope(call: CallbackQuery, state: FSMContext):
    scope = call.data.split(":",1)[1]  # day|slot
    await state.update_data(scope=scope)

    if scope == "slot":
        kb = InlineKeyboardBuilder()
        for s in STD_SLOTS:
            kb.button(text=s, callback_data=f"limit_slot:{s}")
        kb.adjust(3)
        await call.message.edit_text("Выбери слот:", reply_markup=kb.as_markup())
        await state.set_state(AdminLimitsState.choosing_slot)
    else:
        kb = InlineKeyboardBuilder()
        for title, code in ROLE_CODES:
            kb.button(text=title, callback_data=f"limit_role:{code}")
        kb.adjust(2)
        await call.message.edit_text("Выбери роль:", reply_markup=kb.as_markup())
        await state.set_state(AdminLimitsState.choosing_role)
    await call.answer()


@dp.callback_query(AdminLimitsState.choosing_slot, F.data.startswith("limit_slot:"))
async def admin_limits_pick_slot(call: CallbackQuery, state: FSMContext):
    slot = call.data.split(":",1)[1]
    await state.update_data(slot=slot)
    kb = InlineKeyboardBuilder()
    for title, code in ROLE_CODES:
        kb.button(text=title, callback_data=f"limit_role:{code}")
    kb.adjust(2)
    await call.message.edit_text(f"Слот: {slot}\nТеперь выбери роль:", reply_markup=kb.as_markup())
    await state.set_state(AdminLimitsState.choosing_role)
    await call.answer()


@dp.callback_query(AdminLimitsState.choosing_role, F.data.startswith("limit_role:"))
async def admin_limits_pick_role(call: CallbackQuery, state: FSMContext):
    role = call.data.split(":",1)[1]
    await state.update_data(role=role)
    # просим число + показываем кнопку отмены
    await call.message.edit_text("Введи максимальное количество (целое число ≥ 0):")
    await call.message.answer("Можно отменить ввод:", reply_markup=cancel_kb())
    await state.set_state(AdminLimitsState.waiting_for_count)
    await call.answer()


@dp.message(AdminLimitsState.waiting_for_count)
async def admin_limits_set_count(message: types.Message, state: FSMContext):
    if is_cancel(message.text):
        await state.clear()
        await message.answer("❌ Отменено.", reply_markup=menu_keyboard())
        return

    try:
        n = int(message.text.strip())
        if n < 0:
            raise ValueError
    except Exception:
        await message.answer("Нужно целое число ≥ 0. Введи ещё раз:"); return

    data = await state.get_data()
    team_id = data["team_id"]; date_iso = data["date"]
    role = data["role"]; scope = data.get("scope","day"); slot = data.get("slot")

    if scope == "day":
        exist = supabase.table("limits").select("id").eq("team_id", team_id).eq("date", date_iso)\
                 .is_("slot", None).eq("role", role).execute().data
        if exist:
            supabase.table("limits").update({"max_count": n}).eq("id", exist[0]["id"]).execute()
        else:
            supabase.table("limits").insert({"team_id": team_id, "date": date_iso, "slot": None, "role": role, "max_count": n}).execute()
        msg = f"✅ Лимит на день {date_iso} для роли «{role}»: {n}"
    else:
        exist = supabase.table("limits").select("id").eq("team_id", team_id).eq("date", date_iso)\
                 .eq("slot", slot).eq("role", role).execute().data
        if exist:
            supabase.table("limits").update({"max_count": n}).eq("id", exist[0]["id"]).execute()
        else:
            supabase.table("limits").insert({"team_id": team_id, "date": date_iso, "slot": slot, "role": role, "max_count": n}).execute()
        msg = f"✅ Лимит на {date_iso} слот {slot} для роли «{role}»: {n}"

    await message.answer(msg, reply_markup=menu_keyboard())
    await state.clear()


# --- Limits view ---
@dp.callback_query(F.data == "admin_limits_view")
async def admin_limits_view(call: CallbackQuery, state: FSMContext):
    me = supabase.table("users").select("team_id,is_admin,is_owner").eq("telegram_id", call.from_user.id).execute().data
    if not me or not ensure_admin(me[0]):
        await call.answer("Нет доступа", show_alert=True); return
    team_id = me[0]["team_id"]

    week = get_active_week(team_id)
    if not week:
        await call.message.edit_text("Сначала создай активную неделю (меню → 📆 Активная неделя).")
        await call.answer(); return

    days = get_week_dates(week["start_date"], week["end_date"])

    def fmt_one_day_limits(day_iso: str) -> str:
        rows = supabase.table("limits").select("slot,role,max_count").eq("team_id", team_id).eq("date", day_iso).execute().data
        if not rows:
            return "—"
        by_role = {}
        for r in rows:
            role = r.get("role") or "—"
            by_role.setdefault(role, {"day": None, "slots": {}})
            if r["slot"] is None:
                by_role[role]["day"] = r["max_count"]
            else:
                by_role[role]["slots"][r["slot"]] = r["max_count"]
        parts = []
        for role in sorted(by_role.keys()):
            rec = by_role[role]
            chunk = f"{role}: "
            sub = []
            if rec["day"] is not None:
                sub.append(f"день={rec['day']}")
            if rec["slots"]:
                slot_str = ", ".join(f"{s}={cnt}" for s, cnt in sorted(rec["slots"].items()))
                sub.append(slot_str)
            chunk += "; ".join(sub) if sub else "—"
            parts.append(chunk)
        return " | ".join(parts)

    header = f"📊 Лимиты на неделю {week['start_date']} — {week['end_date']}\n"
    msg = header
    sent_any = False
    for d in days:
        line = f"{d['weekday']} {d['date']}: {fmt_one_day_limits(d['date_iso'])}\n"
        if len(msg) + len(line) > 3500:
            await call.message.answer(msg)
            msg = ""
            sent_any = True
        msg += line

    if msg:
        if sent_any:
            await call.message.answer(msg)
        else:
            await call.message.edit_text(msg)

    kb = InlineKeyboardBuilder()
    kb.button(text="⬅️ Назад в админ-меню", callback_data="admin_back")
    await call.message.answer("Готово.", reply_markup=kb.as_markup())
    await call.answer()


@dp.callback_query(F.data == "admin_back")
async def admin_back(call: CallbackQuery, state: FSMContext):
    me = supabase.table("users").select("id,team_id,is_admin,is_owner").eq("telegram_id", call.from_user.id).execute().data
    if not me or not ensure_admin(me[0]):
        await call.answer("Нет доступа", show_alert=True); return
    kb = InlineKeyboardBuilder()
    kb.button(text="📆 Активная неделя", callback_data="admin_week")
    kb.button(text="📈 Лимиты (создать/изменить)", callback_data="admin_limits")
    kb.button(text="👀 Лимиты недели (просмотр)", callback_data="admin_limits_view")
    kb.button(text="🔁 Скопировать лимиты → след. неделя", callback_data="admin_limits_copy_next")
    kb.button(text="📥 Скачать Excel", callback_data="admin_download_excel")
    kb.button(text="👤 Участники", callback_data="admin_members")
    kb.button(text="♻️ Сбросить инвайт-код", callback_data="admin_reset_invite")
    kb.adjust(1)
    await call.message.edit_text("Админ-панель:", reply_markup=kb.as_markup())
    await call.answer()


# --- Limits copy to next week ---
@dp.callback_query(F.data == "admin_limits_copy_next")
async def admin_limits_copy_next(call: CallbackQuery, state: FSMContext):
    me = supabase.table("users").select("team_id,is_admin,is_owner").eq("telegram_id", call.from_user.id).execute().data
    if not me or not ensure_admin(me[0]):
        await call.answer("Нет доступа", show_alert=True); return
    team_id = me[0]["team_id"]

    week = get_active_week(team_id)
    if not week:
        await call.message.edit_text("Сначала создай активную неделю (меню → 📆 Активная неделя).")
        await call.answer(); return

    start = datetime.strptime(week["start_date"], "%Y-%m-%d").date()
    end   = datetime.strptime(week["end_date"], "%Y-%m-%d").date()

    rows = supabase.table("limits").select("date,slot,role,max_count") \
        .eq("team_id", team_id).gte("date", start.isoformat()).lte("date", end.isoformat()).execute().data

    if not rows:
        await call.message.edit_text("На активной неделе нет лимитов для копирования.")
        await call.answer(); return

    inserted = 0
    updated = 0
    for r in rows:
        src_date = datetime.strptime(r["date"], "%Y-%m-%d").date()
        dst_date = (src_date + timedelta(days=7)).isoformat()
        role = r["role"]
        slot = r["slot"]  # может быть None
        max_count = r["max_count"]

        q = supabase.table("limits").select("id").eq("team_id", team_id).eq("date", dst_date).eq("role", role)
        if slot is None:
            q = q.is_("slot", None)
        else:
            q = q.eq("slot", slot)
        exist = q.execute().data

        if exist:
            supabase.table("limits").update({"max_count": max_count}).eq("id", exist[0]["id"]).execute()
            updated += 1
        else:
            supabase.table("limits").insert({
                "team_id": team_id, "date": dst_date, "slot": slot, "role": role, "max_count": max_count
            }).execute()
            inserted += 1

    await call.message.edit_text(
        f"✅ Скопировано лимитов на следующую неделю: добавлено {inserted}, обновлено {updated}."
    )
    await call.answer()


# --- Reset invite code ---
@dp.callback_query(F.data == "admin_reset_invite")
async def admin_reset_invite(call: CallbackQuery, state: FSMContext):
    me = supabase.table("users").select("team_id,is_admin,is_owner").eq("telegram_id", call.from_user.id).execute().data
    if not me or not ensure_admin(me[0]):
        await call.answer("Нет доступа", show_alert=True); return
    team_id = me[0]["team_id"]
    new_code = str(uuid4()).split("-")[0].upper()
    supabase.table("teams").update({"invite_code": new_code}).eq("id", team_id).execute()
    await call.message.edit_text(f"♻️ Новый инвайт-код: <code>{new_code}</code>", parse_mode="HTML")
    await call.answer()


# --- Members: list / card / actions ---

def _member_badges(u: dict) -> str:
    badges = []
    if u.get("is_owner"): badges.append("👑")
    elif u.get("is_admin"): badges.append("🛡️")
    badges.append("🟢" if u.get("is_active", True) else "🔴")
    return "".join(badges)


def _paginate(items, page, size):
    total = len(items)
    start = page * size
    end = start + size
    return items[start:end], total


@dp.callback_query(F.data == "admin_members")
async def admin_members_start(call: CallbackQuery, state: FSMContext):
    me = supabase.table("users").select("team_id,is_admin,is_owner").eq("telegram_id", call.from_user.id).execute().data
    if not me or not ensure_admin(me[0]):
        await call.answer("Нет доступа", show_alert=True); return
    team_id = me[0]["team_id"]

    members = supabase.table("users").select("id,name,role,is_admin,is_owner,is_active") \
        .eq("team_id", team_id).order("name").execute().data

    await state.update_data(members_cache=members)  # кэш на время просмотра
    await _render_members_page(call.message, members, page=0)
    await state.set_state(AdminMembersState.browsing)
    await call.answer()


async def _render_members_page(msg: types.Message, members: list, page: int):
    page_items, total = _paginate(members, page, PAGE_SIZE)
    kb = InlineKeyboardBuilder()
    for u in page_items:
        label = f"{_member_badges(u)} {u['name']} ({u.get('role') or '—'})"
        kb.button(text=label[:64], callback_data=f"member_open:{u['id']}")
    # навигация
    nav = InlineKeyboardBuilder()
    has_prev = page > 0
    has_next = (page + 1) * PAGE_SIZE < total
    if has_prev:
        nav.button(text="⬅️ Назад", callback_data=f"members_page:{page-1}")
    if has_next:
        nav.button(text="Вперёд ➡️", callback_data=f"members_page:{page+1}")
    nav.adjust(2)
    kb.adjust(1)
    text = f"👤 Участники (стр. {page+1})"
    await msg.edit_text(text, reply_markup=kb.as_markup())
    if has_prev or has_next:
        await msg.answer("Навигация:", reply_markup=nav.as_markup())


@dp.callback_query(AdminMembersState.browsing, F.data.startswith("members_page:"))
async def members_page_nav(call: CallbackQuery, state: FSMContext):
    page = int(call.data.split(":")[1])
    data = await state.get_data()
    members = data.get("members_cache", [])
    await _render_members_page(call.message, members, page)
    await call.answer()


@dp.callback_query(F.data.startswith("member_open:"))
async def member_open(call: CallbackQuery, state: FSMContext):
    member_id = call.data.split(":")[1]
    me = supabase.table("users").select("team_id,is_admin,is_owner,id").eq("telegram_id", call.from_user.id).execute().data[0]
    if not ensure_admin(me):
        await call.answer("Нет доступа", show_alert=True); return
    team_id = me["team_id"]

    rows = supabase.table("users").select("id,name,role,is_admin,is_owner,is_active,telegram_id") \
        .eq("id", member_id).eq("team_id", team_id).execute().data
    if not rows:
        await call.answer("Пользователь не найден.", show_alert=True); return
    u = rows[0]

    text = (
        f"{_member_badges(u)} <b>{u['name']}</b>\n"
        f"Роль: <code>{u.get('role') or '—'}</code>\n"
        f"Статус: {'Активен' if u.get('is_active', True) else 'Отключён'}\n"
        f"Права: {'Владелец' if u.get('is_owner') else ('Админ' if u.get('is_admin') else 'Сотрудник')}"
    )

    kb = InlineKeyboardBuilder()
    # смена роли
    for title, code in ROLE_CODES:
        kb.button(text=title, callback_data=f"member_setrole:{u['id']}:{code}")
    kb.adjust(3)

    actions = InlineKeyboardBuilder()
    # нельзя понижать владельца по кнопке "админ"
    if not u.get("is_owner"):
        actions.button(
            text=("Снять админа" if u.get("is_admin") else "Сделать админом"),
            callback_data=f"member_admin_toggle:{u['id']}"
        )
    actions.button(
        text=("Отключить" if u.get("is_active", True) else "Восстановить"),
        callback_data=f"member_toggle_active:{u['id']}"
    )
    if me["id"] != u["id"]:
        actions.button(text="Удалить из команды", callback_data=f"member_remove:{u['id']}")
    actions.button(text="↩️ К списку", callback_data="admin_members")
    actions.adjust(2)

    await call.message.edit_text(text, parse_mode="HTML", reply_markup=kb.as_markup())
    await call.message.answer("Действия:", reply_markup=actions.as_markup())
    await state.set_state(AdminMembersState.member_card)
    await call.answer()


@dp.callback_query(F.data.startswith("member_setrole:"))
async def member_setrole(call: CallbackQuery, state: FSMContext):
    _, user_id, role = call.data.split(":")
    me = supabase.table("users").select("team_id,is_admin,is_owner").eq("telegram_id", call.from_user.id).execute().data[0]
    if not ensure_admin(me):
        await call.answer("Нет доступа", show_alert=True); return
    supabase.table("users").update({"role": role}).eq("id", user_id).eq("team_id", me["team_id"]).execute()
    await call.answer("Роль обновлена")
    # перерисуем карточку
    await member_open(call, state)


@dp.callback_query(F.data.startswith("member_admin_toggle:"))
async def member_admin_toggle(call: CallbackQuery, state: FSMContext):
    user_id = call.data.split(":")[1]
    me = supabase.table("users").select("team_id,is_admin,is_owner,id").eq("telegram_id", call.from_user.id).execute().data[0]
    if not ensure_admin(me):
        await call.answer("Нет доступа", show_alert=True); return
    rows = supabase.table("users").select("is_admin,is_owner").eq("id", user_id).eq("team_id", me["team_id"]).execute().data
    if not rows:
        await call.answer("Не найдено", show_alert=True); return
    u = rows[0]
    if u.get("is_owner"):
        await call.answer("Нельзя изменять права владельца.", show_alert=True); return
    supabase.table("users").update({"is_admin": not u.get("is_admin", False)}).eq("id", user_id).execute()
    await call.answer("Готово")
    await member_open(call, state)


@dp.callback_query(F.data.startswith("member_toggle_active:"))
async def member_toggle_active(call: CallbackQuery, state: FSMContext):
    user_id = call.data.split(":")[1]
    me = supabase.table("users").select("team_id,is_admin,is_owner").eq("telegram_id", call.from_user.id).execute().data[0]
    if not ensure_admin(me):
        await call.answer("Нет доступа", show_alert=True); return
    rows = supabase.table("users").select("is_active").eq("id", user_id).eq("team_id", me["team_id"]).execute().data
    if not rows:
        await call.answer("Не найдено", show_alert=True); return
    curr = rows[0].get("is_active", True)
    if curr:
        supabase.table("users").update({"is_active": False, "left_at": now_iso_z(), "is_admin": False}).eq("id", user_id).execute()
    else:
        supabase.table("users").update({"is_active": True, "left_at": None}).eq("id", user_id).execute()
    await call.answer("Статус изменён")
    await member_open(call, state)


@dp.callback_query(F.data.startswith("member_remove:"))
async def member_remove(call: CallbackQuery, state: FSMContext):
    user_id = call.data.split(":")[1]
    me = supabase.table("users").select("team_id,is_admin,is_owner,id").eq("telegram_id", call.from_user.id).execute().data[0]
    if not ensure_admin(me):
        await call.answer("Нет доступа", show_alert=True); return
    if me["id"] == user_id:
        await call.answer("Нельзя удалить самого себя.", show_alert=True); return
    # Жёсткое удаление из команды: team_id=null, снимаем админа и деактивируем
    supabase.table("users").update({"team_id": None, "is_admin": False, "is_active": False}).eq("id", user_id).eq("team_id", me["team_id"]).execute()
    await call.answer("Пользователь удалён из команды")
    await admin_members_start(call, state)


# --- EXPORT BUTTON HANDLER ---
@dp.callback_query(F.data == "admin_download_excel")
async def admin_download_excel(call: CallbackQuery, state: FSMContext):
    me = supabase.table("users").select("team_id,is_admin,is_owner").eq("telegram_id", call.from_user.id).execute().data
    if not me or not ensure_admin(me[0]):
        await call.answer("Нет доступа", show_alert=True); return

    team_id = me[0]["team_id"]
    # Проверим наличие активной недели
    if not get_active_week(team_id):
        await call.answer("Нет активной недели", show_alert=True); return

    await call.message.answer("Готовлю файл…")
    try:
        filename, bio = export_schedule_excel_for_team(team_id)
    except Exception as e:
        await call.message.answer(f"Ошибка экспорта: {e}")
        await call.answer()
        return

    document = types.BufferedInputFile(bio.getvalue(), filename=filename)
    await call.message.answer_document(document)
    await call.answer()


# ---------------- GLOBAL CANCEL ----------------
# Ловит "Отмена", "❌ Отмена" и т.п. в ЛЮБОМ состоянии (дополнительно к ранней проверке в state-хэндлерах)
@dp.message(F.text.regexp(r"(?i)отмена"))
async def cancel_text(message: types.Message, state: FSMContext):
    await state.clear()
    await message.answer("❌ Отменено.", reply_markup=menu_keyboard())


# ---------------- RUN ----------------
if __name__ == "__main__":
    import asyncio, logging, sys
    logging.basicConfig(level=logging.INFO)
    print("Бот стартует... запускаю polling")
    if sys.platform == "win32":
        asyncio.set_event_loop_policy(asyncio.WindowsSelectorEventLoopPolicy())
    try:
        dp.run_polling(bot)
    finally:
        print("Polling завершён")
