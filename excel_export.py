# excel_export.py
import io
import pandas as pd
from datetime import datetime, timedelta
from dateutil import tz
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Конфиг оформления
GROUP_FILL = "FFA500"   # оранжевый для заголовков подгрупп
WEEKEND_FILL = "FFF2CC" # мягкое выделение выходных колонок
HEADER_FILL = "EFEFEF"  # шапка
BORDER = Side(style="thin", color="DDDDDD")

ROLE_ORDER = ["ОФИЦИАНТ", "ХОСТ", "БАРМЕН", "РАНЕР", "АДМИН"]  # порядок блоков

def _monday(d: datetime) -> datetime:
    return d - timedelta(days=d.weekday())

def _dates_of_week(monday: datetime):
    return [monday + timedelta(days=i) for i in range(7)]

def _fmt_dt(d: datetime) -> str:
    return d.strftime("%d.%m")

def _fetch_limits(supabase, team_id=None):
    # ожидается таблица limits с колонками: weekday(0-6), role, max_count, team_id (опционально)
    q = supabase.table("limits").select("*")
    if team_id:
        q = q.eq("team_id", team_id)
    data = q.execute().data or []
    limits = {}
    for row in data:
        weekday = int(row.get("weekday", 0))
        role = (row.get("role") or "").upper()
        max_count = int(row.get("max_count", 0))
        limits.setdefault(weekday, {})[role] = max_count
    return limits

def _fetch_users_by_role(supabase, team_id=None):
    # ожидается таблица users с колонками: id, full_name, role, team_id (опц)
    q = supabase.table("users").select("id,full_name,role")
    if team_id:
        q = q.eq("team_id", team_id)
    data = q.execute().data or []
    users_by_role = {}
    for u in data:
        role = (u.get("role") or "").upper()
        users_by_role.setdefault(role, []).append(u)
    # сортировка по имени
    for r in users_by_role:
        users_by_role[r].sort(key=lambda x: (x.get("full_name") or "").lower())
    return users_by_role

def _fetch_shifts(supabase, start_iso: str, end_iso: str, team_id=None):
    # ожидается shifts с: user_id, date (ISO), slot, role_override (опц), team_id (опц)
    q = (supabase.table("shifts")
         .select("user_id,date,slot,role_override")
         .gte("date", start_iso).lt("date", end_iso))
    if team_id:
        q = q.eq("team_id", team_id)
    data = q.execute().data or []
    return data

def _build_dataframe(users_by_role, shifts, week_days):
    """
    Создаёт список блоков DataFrame по каждой роли.
    Колонки: ФИО + 7 дней (пн-вс), ячейки = слот/роль (если переопределена).
    """
    # Индекс быстрых lookup
    # map[(user_id, yyyy-mm-dd)] -> {"slot": "...", "role_override": "..."}
    shift_map = {}
    for s in shifts:
        key = (s["user_id"], s["date"][:10])
        shift_map[key] = {"slot": s.get("slot"), "role_override": s.get("role_override")}

    blocks = []  # [(role, df)]
    for role in ROLE_ORDER:
        users = users_by_role.get(role, [])
        if not users:
            continue
        rows = []
        for u in users:
            row = {"Сотрудник": u["full_name"]}
            for d in week_days:
                key = (u["id"], d.strftime("%Y-%m-%d"))
                cell = ""
                if key in shift_map:
                    cell_slot = shift_map[key].get("slot") or ""
                    cell_role = (shift_map[key].get("role_override") or "").upper()
                    cell = cell_slot
                    if cell_role and cell_role != role:
                        cell = f"{cell_slot} [{cell_role}]"
                row[_fmt_dt(d)] = cell
            rows.append(row)
        df = pd.DataFrame(rows, columns=["Сотрудник"] + [_fmt_dt(d) for d in week_days])
        blocks.append((role, df))
    return blocks

def _apply_styles(ws, start_row, end_row, start_col, end_col, weekend_cols):
    # Границы и выравнивание
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
            cell.border = Border(top=BORDER, bottom=BORDER, left=BORDER, right=BORDER)

    # Выходные дни (сб=5, вс=6)
    for c in weekend_cols:
        for r in range(start_row + 1, end_row + 1):  # не красим заголовок группы
            cell = ws.cell(row=r, column=c)
            cell.fill = PatternFill("solid", fgColor=WEEKEND_FILL)

def _auto_width(ws):
    for col in range(1, ws.max_column + 1):
        max_len = 0
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(col)].width = min(max(12, max_len + 2), 38)

def export_schedule_excel(supabase, tz_name="Europe/Moscow", week_monday: datetime|None=None, team_id=None):
    """
    Возвращает BytesIO с Excel-файлом расписания на неделю.
    """
    tzinfo = tz.gettz(tz_name)
    now = datetime.now(tzinfo)
    monday = _monday(now) if week_monday is None else week_monday
    days = _dates_of_week(monday)

    # Загружаем данные
    users_by_role = _fetch_users_by_role(supabase, team_id)
    shifts = _fetch_shifts(
        supabase,
        start_iso=days[0].strftime("%Y-%m-%d"),
        end_iso=(days[-1] + timedelta(days=1)).strftime("%Y-%m-%d"),
        team_id=team_id
    )
    limits = _fetch_limits(supabase, team_id)

    # Сборка датафреймов по ролям
    blocks = _build_dataframe(users_by_role, shifts, days)

    # Если нет данных — всё равно создадим пустую страницу с датами
    if not blocks:
        empty = pd.DataFrame(columns=["Сотрудник"] + [_fmt_dt(d) for d in days])
        blocks = [("ОФИЦИАНТ", empty)]

    # Пишем в Excel
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # Один лист "График"
        sheet_name = "График"
        start_row = 1

        # Сначала — сводка по лимитам вверху
        # Структура: День недели | Роль | Лимит
        lim_rows = []
        wd_ru = ["Пн","Вт","Ср","Чт","Пт","Сб","Вс"]
        for wd in range(7):
            day_label = f"{wd_ru[wd]} {_fmt_dt(days[wd])}"
            roles_map = limits.get(wd, {})
            if not roles_map:
                lim_rows.append([day_label, "—", "—"])
                continue
            for role, lim in roles_map.items():
                lim_rows.append([day_label, role, lim])
        lim_df = pd.DataFrame(lim_rows, columns=["День", "Роль", "Лимит"])
        lim_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=start_row-1)
        start_row += (len(lim_rows) + 3 if lim_rows else 3)  # отступ после сводки

        # Далее — блоки по ролям
        first_block_row = start_row
        for role, df in blocks:
            df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=start_row-1)
            start_row += (len(df) + 3)  # отступ между блоками

        # Стилей добавляем после записи
        wb = writer.book
        ws = wb[sheet_name]

        # Шапка сводки лимитов
        for c in range(1, 4):
            cell = ws.cell(row=1, column=c)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Подпись и оформление блоков по ролям
        cur_row = first_block_row
        weekend_cols = []  # вычислим по первой таблице (Сб/Вс)
        for role, df in blocks:
            n_rows = len(df)
            n_cols = df.shape[1]  # "Сотрудник" + 7 дней
            start_c = 1
            end_c = n_cols

            # Заголовок подгруппы — отдельная строка перед таблицей
            ws.insert_rows(cur_row, 1)
            ws.merge_cells(start_row=cur_row, start_column=start_c, end_row=cur_row, end_column=end_c)
            title_cell = ws.cell(row=cur_row, column=start_c)
            title_cell.value = role.title()
            title_cell.font = Font(bold=True)
            title_cell.alignment = Alignment(horizontal="left", vertical="center")
            title_cell.fill = PatternFill("solid", fgColor=GROUP_FILL)

            # Заголовок колонок таблицы (имена дней) — делаем жирным и серым
            header_row = cur_row + 1
            for c in range(start_c, end_c + 1):
                hcell = ws.cell(row=header_row, column=c)
                hcell.font = Font(bold=True)
                hcell.fill = PatternFill("solid", fgColor=HEADER_FILL)
                hcell.alignment = Alignment(horizontal="center", vertical="center")

            # Вычислим индексы колонок Сб/Вс (по заголовкам)
            if not weekend_cols:
                # Ищем по тексту "дд.мм" и сопоставляем с днями недели
                day_cols = {}
                for c in range(1, end_c + 1):
                    val = ws.cell(row=header_row, column=c).value
                    if isinstance(val, str) and len(val) == 5 and val[2] == ".":
                        day_cols[c] = val
                # Сопоставим позиции дат с исходными days
                txt_to_wd = {_fmt_dt(days[i]): i for i in range(7)}
                for c, txt in day_cols.items():
                    wd = txt_to_wd.get(txt)
                    if wd in (5, 6):  # 5=Сб, 6=Вс
                        weekend_cols.append(c)

            # Сетка, выравнивание, выходные
            table_start = header_row
            table_end = header_row + n_rows
            _apply_styles(ws, table_start, table_end, start_c, end_c, weekend_cols)

            # Имена сотрудников — делаем «белыми» только в логике UI; в Excel оставим стандартный чёрный шрифт для печати.
            # Если очень нужно «белые», можно поставить цвет FFFFFF, но в печати это будет плохо.
            # Поэтому оставляем чёрный, как вы и просили для читаемости.

            cur_row = table_end + 2  # отступ перед следующим блоком

        # Авто-ширина
        _auto_width(ws)

        # Титульная ячейка сверху
        ws.insert_rows(1, 1)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ws.max_column)
        title = ws.cell(row=1, column=1)
        title.value = f"График на неделю: {days[0].strftime('%d.%m.%Y')} — {days[-1].strftime('%d.%m.%Y')}"
        title.font = Font(bold=True)
        title.alignment = Alignment(horizontal="left", vertical="center")

    output.seek(0)
    filename = f"grafik_{days[0].strftime('%Y%m%d')}_{days[-1].strftime('%Y%m%d')}.xlsx"
    return filename, output
